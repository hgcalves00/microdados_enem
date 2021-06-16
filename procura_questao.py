import re   
import fitz
from openpyxl import load_workbook

#a planilha estará disponibilizada na pasta!
planilha_p = load_workbook(filename = 'microdados.xlsx')
planilha = planilha_p['PROVAS']

#procurar dentro do pdf de acordo com o ano
def procuraPDF(ano, procura): 
    pdf = fitz.open('caderno_enem_{}.pdf'.format(ano))
    paginas = pdf.pageCount
    listinha = []
    for i in range(0, paginas):
        pagina = pdf.loadPage(i)
        textop = pagina.getTextPage()
        texto = textop.extractText()
        if procura in texto:
            x = texto.index(procura)
            listinha = re.findall("Questao \d", texto)
            if listinha == []:
                listinha = re.findall("Questao \d\d", texto)
                if listinha == []:
                    listinha = re.findall("Questao \d\d\d", texto)
                    if listinha == []:
                        listinha.append(texto[x-18:x-1])
                        if listinha == []:
                            print('não encontrei! :( tente escrever exatamente o que está no texto')
    lista_do_numero = re.findall('\d', listinha[0])
    numero = ''
    if lista_do_numero[0] == 0:
        numero = numero + lista_do_numero[1]
    else:
        if len(lista_do_numero) == 1:
            numero = lista_do_numero[0]
        elif len(lista_do_numero) == 2:
            numero = lista_do_numero[0] + lista_do_numero[1]
        elif len(lista_do_numero) == 3:
            numero = lista_do_numero[0] + lista_do_numero[1] + lista_do_numero[2]
    numero = int(numero)
    infos = [ano, numero]
    return infos

#bate a planilha com o que foi achado na procura dentro do pdf
def acha_na_planilha(ano, pesquisa):
    dados = procuraPDF(ano, pesquisa)
    for i in range(1, planilha.max_row + 1):
        if planilha[i][1].value == dados[0] and planilha[i][2].value == dados[1]:
           infos = [planilha[i][0].value, planilha[i][1].value, planilha[i][2].value, planilha[i][3].value, planilha[i][4].value, planilha[i][5].value, planilha[i][6].value]
    return infos

#exemplo da prova de 2009 (funciona)
print(acha_na_planilha(2009, 'A atmosfera terrestre é composta'))