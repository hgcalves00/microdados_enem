from openpyxl import load_workbook

planilha_p = load_workbook(filename = r'microdados.xlsx')
planilha_1 = planilha_p['PROVA_2015']

def texto_da_prova():
    with open(r'2015.txt', 'r', encoding='utf8') as txt:
        linhas = txt.readlines()
    return linhas

indices_numeros = [2, 12, 22, 34, 46, 56, 66, 78, 90, 101, 111, 120, 131, 142, 152, 169, 183, 197, 208, 219, 229, 239, 251, 263, 274, 285, 297, 310, 320, 331, 343, 355, 390, 401, 413, 423, 433, 443, 454, 464, 474, 485, 497, 509, 519, 529, 550, 561, 574, 584, 594, 604, 614, 628, 639, 649, 661, 671, 681, 691, 701, 713, 724, 734, 744, 757, 768, 778, 788, 798, 808, 819, 831, 841, 851, 861, 871, 883, 895, 907, 921, 931, 943, 953, 963, 974, 984, 996, 1011, 1023, 1033, 1043, 1060, 1071, 1083, 1096, 1107, 1121, 1141, 1150, 1166, 1176, 1187, 1198, 1222, 1235, 1246, 1256, 1266, 1279, 1293, 1304, 1331, 1348, 1359, 1370, 1383, 1395, 1424, 1436, 1470, 1479, 1492, 1503, 1521, 1548, 1560, 1570, 1598, 1612, 1626, 1636, 1650, 1664, 1679, 1707, 1734, 1763, 1787, 1801, 1816, 1827, 1839, 1849, 1863, 1877, 1891, 1901, 1914, 1924, 1934, 1944, 1958, 1970, 1980, 1992, 2004, 2014, 2024, 2036, 2046, 2056, 2072, 2090, 2100, 2110, 2122, 2136, 2146, 2156, 2168, 2178, 2190, 2202, 2212, 2224, 2240, 2258, 2272, 2282, 2298, 2310, 2320, 2334, 2346]

def pesquisa(pesquisou):
    n = len(indices_numeros)
    trecho_string = ''
    for i in range(n-1):
        trecho = texto_da_prova()[indices_numeros[i]:indices_numeros[i+1]]
        for a in trecho:
            if pesquisou in a:
                ponto = a.find('.')
                numero = a[0:ponto]
                for b in trecho:
                    trecho_string += b
    if numero == '':
        ponto = trecho_string.find('.')
        numero= trecho_string[52:ponto]
    return numero

def planilha(pesquisou):
    numero = pesquisa(pesquisou)
    print(numero)
    for i in range(2, planilha_1.max_row + 1):
        if planilha_1[i][2].value == numero:
            print(planilha_1[i][0].value)
            print(planilha_1[i][1].value)
            print(planilha_1[i][2].value)
            print(planilha_1[i][3].value)
            print(planilha_1[i][4].value)
            print(planilha_1[i][5].value)
            print(planilha_1[i][6].value)
